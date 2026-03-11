import streamlit as st
import openpyxl
import networkx as nx
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import re
from collections import defaultdict
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Excel Model Auditor", layout="wide")
st.title("Excel Model Auditor")
st.caption("Upload a workbook to visualize its dependency graph and detect structural issues.")

# --- Sidebar ---
with st.sidebar:
    st.header("Settings")
    uploaded_file = st.file_uploader("Upload Excel workbook", type=["xlsx"])
    input_sheets = st.text_input(
        "Input sheet names (comma-separated)",
        value="Inputs",
        help="Cells in these sheets are expected to hold constants — they won't be flagged as hardcoded.",
    )
    input_sheet_list = [s.strip() for s in input_sheets.split(",") if s.strip()]
    view_mode = st.radio(
        "Graph view",
        ["Sheet overview", "Cell detail"],
        help="Sheet overview is readable for any size model. Cell detail works best for small workbooks.",
    )

if uploaded_file is None:
    st.info("Upload an Excel file in the sidebar to get started.")
    st.stop()

# --- Helpers ---
def get_formula_str(value):
    """Return the formula as a plain string.
    openpyxl returns dynamic array formulas (FILTER, UNIQUE, SORT, XLOOKUP, etc.)
    as ArrayFormula objects rather than strings — unwrap them here."""
    if hasattr(value, "text"):
        return value.text or ""
    return value or ""

def expand_range(sheet, start, end):
    col1, row1 = coordinate_from_string(start)
    col2, row2 = coordinate_from_string(end)
    col1_i = column_index_from_string(col1)
    col2_i = column_index_from_string(col2)
    cells = []
    for col in range(col1_i, col2_i + 1):
        for row in range(row1, row2 + 1):
            cells.append(f"{sheet}!{get_column_letter(col)}{row}")
    return cells

SHEET_COLORS = ["lightgreen", "lightskyblue", "salmon", "plum", "peachpuff", "lightcyan", "wheat", "lavender"]

def sheet_color(sheet):
    idx = sheet_names.index(sheet) if sheet in sheet_names else -1
    return SHEET_COLORS[idx % len(SHEET_COLORS)]

def cell_node_color(node, hardcoded_cells):
    if node in hardcoded_cells:
        return "gold"
    sheet = node.split("!")[0].strip("'")
    return sheet_color(sheet)

# --- Load workbook ---
wb = openpyxl.load_workbook(uploaded_file, data_only=False)
sheet_names = wb.sheetnames

st.subheader("Sheets detected")
st.write(", ".join(sheet_names))

# --- Preprocess: expand named ranges in formulas ---
named_ranges_map = {}
for name in wb.defined_names:
    defn = wb.defined_names[name]
    try:
        parts = []
        for sheet_name, cell_range in defn.destinations:
            clean = cell_range.replace("$", "")
            sheet_ref = f"'{sheet_name}'" if " " in sheet_name else sheet_name
            parts.append(f"{sheet_ref}!{clean}")
        named_ranges_map[name] = ",".join(parts)
    except Exception:
        pass

if named_ranges_map:
    sorted_names = sorted(named_ranges_map, key=len, reverse=True)
    for sheet in sheet_names:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" and cell.value:
                    formula = get_formula_str(cell.value)
                    for name in sorted_names:
                        pattern = re.compile(r"\b" + re.escape(name) + r"\b", re.IGNORECASE)
                        formula = pattern.sub(named_ranges_map[name], formula)
                    cell.value = formula

# --- Detect hardcoded constants ---
hardcoded_cells = set()
for sheet in sheet_names:
    if sheet in input_sheet_list:
        continue
    ws = wb[sheet]
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "n" and cell.value is not None:
                hardcoded_cells.add(f"{sheet}!{cell.coordinate}")

# --- Build cell-level dependency graph ---
G = nx.DiGraph()

# Sheet name: either quoted (allowing spaces) e.g. 'Balance Sheet', or plain word e.g. FinStat
_sheet = r"(?:'[^']+'|[A-Za-z_][A-Za-z_0-9]*)"
# Individual cell ref:  Sheet!A1  or  'Sheet Name'!$A$1
cell_pattern    = rf"{_sheet}!\$?[A-Za-z]{{1,3}}\$?\d+"
# Range ref:            Sheet!A1:B10
range_pattern   = rf"({_sheet})!([A-Za-z]{{1,3}}\d+):([A-Za-z]{{1,3}}\d+)"

for sheet in sheet_names:
    ws = wb[sheet]
    for row in ws.iter_rows():
        for cell in row:
            location = f"{sheet}!{cell.coordinate}"
            if cell.data_type == "f":
                formula = get_formula_str(cell.value)
                for r in re.findall(range_pattern, formula):
                    rng_sheet, start, end = r
                    rng_sheet = rng_sheet.strip("'")
                    for ref in expand_range(rng_sheet, start, end):
                        G.add_edge(ref, location)
                for ref in re.findall(cell_pattern, formula):
                    ref = ref.replace("$", "").strip("'")
                    # Re-attach clean sheet name: strip quotes then rebuild
                    parts = ref.split("!")
                    if len(parts) == 2:
                        G.add_edge(ref, location)

for cell in hardcoded_cells:
    if cell not in G:
        G.add_node(cell)

# --- Build sheet-level summary graph ---
SG = nx.DiGraph()
sheet_edge_counts = defaultdict(int)
for src, dst in G.edges:
    src_sheet = src.split("!")[0].strip("'")
    dst_sheet = dst.split("!")[0].strip("'")
    if src_sheet != dst_sheet:
        sheet_edge_counts[(src_sheet, dst_sheet)] += 1

for sheet in sheet_names:
    SG.add_node(sheet)
for (src_sheet, dst_sheet), count in sheet_edge_counts.items():
    SG.add_edge(src_sheet, dst_sheet, weight=count)

# --- Draw graph ---
st.subheader("Dependency graph")

if view_mode == "Sheet overview":
    # Circular layout for sheets — always readable
    pos = nx.circular_layout(SG)
    colors = [sheet_color(n) for n in SG.nodes]
    edge_labels = {(u, v): d["weight"] for u, v, d in SG.edges(data=True)}
    node_sizes = [1500 + SG.degree(n) * 300 for n in SG.nodes]

    fig, ax = plt.subplots(figsize=(10, 8))
    nx.draw_networkx_nodes(SG, pos, ax=ax, node_color=colors, node_size=node_sizes)
    nx.draw_networkx_labels(SG, pos, ax=ax, font_size=9, font_weight="bold")
    nx.draw_networkx_edges(SG, pos, ax=ax, arrows=True,
                           arrowstyle="-|>", arrowsize=20,
                           edge_color="gray", width=1.5,
                           connectionstyle="arc3,rad=0.1")
    nx.draw_networkx_edge_labels(SG, pos, edge_labels=edge_labels, ax=ax,
                                 font_size=7, label_pos=0.3)
    ax.set_title("Sheet-level dependency overview\n(edge numbers = cross-sheet formula references)")
    ax.axis("off")

    legend = [mpatches.Patch(color=sheet_color(s), label=s) for s in sheet_names]
    ax.legend(handles=legend, loc="upper left", fontsize=8)

else:
    # Cell-level — warn if large
    if len(G.nodes) > 300:
        st.warning(f"This workbook has {len(G.nodes)} nodes. The cell-level graph will be dense — sheet overview is recommended.")

    layers = {name: i for i, name in enumerate(sheet_names)}
    pos = {}
    y_positions = {}
    for node in sorted(G.nodes):
        sheet = node.split("!")[0].strip("'")
        layer = layers.get(sheet, len(sheet_names))
        y_positions.setdefault(sheet, 0)
        pos[node] = (layer, -y_positions[sheet])
        y_positions[sheet] += 1

    colors = [cell_node_color(n, hardcoded_cells) for n in G.nodes]
    sizes = [400 + len(nx.descendants(G, n)) * 100 for n in G.nodes]

    fig, ax = plt.subplots(figsize=(14, 8))
    nx.draw(G, pos, ax=ax, with_labels=True, node_color=colors,
            node_size=sizes, font_size=7, arrows=True)
    ax.set_title("Cell-level dependency graph")

    legend = [mpatches.Patch(color="gold", label="Hardcoded constant")]
    for s in sheet_names:
        legend.append(mpatches.Patch(color=sheet_color(s), label=s))
    ax.legend(handles=legend, loc="upper left", fontsize=8)

st.pyplot(fig)
plt.close(fig)

# --- Metrics ---
col1, col2, col3, col4 = st.columns(4)
col1.metric("Sheets", len(sheet_names))
col2.metric("Formula cells", len(G.nodes))
col3.metric("Dependencies", len(G.edges))
col4.metric("Hardcoded constants", len(hardcoded_cells))

# --- Detail tables ---
col_a, col_b = st.columns(2)

with col_a:
    st.subheader("Hardcoded constants")
    st.caption("Numeric values outside input sheets — should these be in Inputs?")
    if hardcoded_cells:
        st.dataframe({"Cell": sorted(hardcoded_cells)}, use_container_width=True)
    else:
        st.success("None found.")

with col_b:
    st.subheader("Orphan cells")
    st.caption("Cells with no connections — overwritten formulas, unused assumptions, or abandoned calculations.")
    orphans = [n for n in G.nodes if G.degree(n) == 0]
    if orphans:
        st.dataframe({"Cell": sorted(orphans)}, use_container_width=True)
    else:
        st.success("None found.")

# --- Cross-sheet reference table ---
with st.expander("Cross-sheet reference counts"):
    rows = [{"From": u, "To": v, "References": c}
            for (u, v), c in sorted(sheet_edge_counts.items(), key=lambda x: -x[1])]
    if rows:
        st.dataframe(rows, use_container_width=True)
    else:
        st.info("No cross-sheet references found.")

if named_ranges_map:
    with st.expander("Named ranges resolved"):
        st.dataframe(
            {"Name": list(named_ranges_map.keys()), "Resolves to": list(named_ranges_map.values())},
            use_container_width=True,
        )
